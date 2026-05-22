#!/usr/bin/env python3
"""
convert_xlsx_to_csv.py — Convert XLSX/XLS files to CSV with integrity verification.

Modes:
  --single-sheet   For files with one sheet: foo.xlsx -> foo.csv
  --multi-sheet    For files with multiple sheets: foo.xlsx -> foo/<sheet>.csv
  --auto           Pick per-file based on sheet count (default)

Verification:
  Re-parses the written CSV and compares cell-by-cell to the original
  XLSX parse. Catches encoding loss, dtype coercion, row count drift.

Originals are NOT deleted by this script. Remove them in a separate
explicit step after batch verification + commit.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

try:
    import xlrd

    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


def normalize_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v)


def read_xlsx_sheet(
    path: Path, sheet_name: str | None = None
) -> tuple[str, list[list[str]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([normalize_cell(c) for c in row])
    wb.close()
    while rows and not any(c.strip() for c in rows[-1]):
        rows.pop()
    return sheet_name, rows


def read_xls_sheet(
    path: Path, sheet_name: str | None = None
) -> tuple[str, list[list[str]]]:
    if not HAS_XLRD:
        raise RuntimeError("xlrd not installed")
    book = xlrd.open_workbook(path)
    sh = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(0)
    sheet_name = sh.name
    rows = []
    for r in range(sh.nrows):
        row = []
        for c in range(sh.ncols):
            cell = sh.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append("")
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                v = cell.value
                row.append(str(int(v)) if v == int(v) else repr(v))
            elif cell.ctype == xlrd.XL_CELL_DATE:
                row.append(str(xlrd.xldate_as_tuple(cell.value, book.datemode)))
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                row.append("TRUE" if cell.value else "FALSE")
            else:
                row.append(str(cell.value))
        rows.append(row)
    while rows and not any(c.strip() for c in rows[-1]):
        rows.pop()
    return sheet_name, rows


def write_csv(rows: list[list[str]], out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, quoting=csv.QUOTE_MINIMAL)
        for row in rows:
            w.writerow(row)


def read_csv_back(path: Path) -> list[list[str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        rows = [list(r) for r in reader]
    while rows and not any(c.strip() for c in rows[-1]):
        rows.pop()
    return rows


def verify(original: list[list[str]], written: list[list[str]]) -> tuple[bool, str]:
    if len(original) != len(written):
        return False, f"row count: orig={len(original)} written={len(written)}"
    for i, (a, b) in enumerate(zip(original, written)):
        max_cols = max(len(a), len(b))
        ap = a + [""] * (max_cols - len(a))
        bp = b + [""] * (max_cols - len(b))
        if [c.rstrip() for c in ap] != [c.rstrip() for c in bp]:
            for j in range(max_cols):
                ca, cb = ap[j], bp[j]
                if ca.rstrip() != cb.rstrip():
                    return False, f"row {i + 1} col {j + 1}: {ca!r} != {cb!r}"
    return True, "OK"


def detect_sheet_count(path: Path) -> int:
    if path.suffix.lower() == ".xls":
        return len(xlrd.open_workbook(path).sheet_names())
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    n = len(wb.sheetnames)
    wb.close()
    return n


def convert_one(xlsx_path: Path, mode: str) -> list[tuple[Path, str, str]]:
    is_xls = xlsx_path.suffix.lower() == ".xls"
    reader = read_xls_sheet if is_xls else read_xlsx_sheet
    results: list[tuple[Path, str, str]] = []

    if is_xls:
        sheets = xlrd.open_workbook(xlsx_path).sheet_names()
    else:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()

    if mode == "single-sheet":
        if len(sheets) != 1:
            return [(xlsx_path, "SKIP", f"multi-sheet ({len(sheets)})")]
        name, rows = reader(xlsx_path)
        out = xlsx_path.with_suffix(".csv")
        write_csv(rows, out)
        written = read_csv_back(out)
        ok, msg = verify(rows, written)
        cols = len(rows[0]) if rows else 0
        return [
            (
                out,
                "OK" if ok else "FAIL",
                f"{len(rows)} rows × {cols} cols" if ok else msg,
            )
        ]

    # multi-sheet
    out_dir = xlsx_path.with_suffix("")
    for sn in sheets:
        name, rows = reader(xlsx_path, sn)
        safe = (
            re.sub(r"[<>:\"/\\|?*\x00-\x1f]+", "_", sn).strip().strip(".")
            or f"sheet_{sheets.index(sn) + 1}"
        )
        out = out_dir / f"{safe}.csv"
        write_csv(rows, out)
        written = read_csv_back(out)
        ok, msg = verify(rows, written)
        cols = len(rows[0]) if rows else 0
        results.append(
            (
                out,
                "OK" if ok else "FAIL",
                f"sheet={sn}: {len(rows)} rows × {cols} cols"
                if ok
                else f"sheet={sn}: {msg}",
            )
        )
    return results


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="+")
    ap.add_argument(
        "--mode", choices=["single-sheet", "multi-sheet", "auto"], default="auto"
    )
    args = ap.parse_args()

    base = Path.cwd()
    all_ok = True
    for f in args.files:
        path = Path(f).resolve()
        if not path.exists():
            print(f"  {f}: NOT FOUND")
            all_ok = False
            continue
        mode = args.mode
        if mode == "auto":
            n = detect_sheet_count(path)
            mode = "single-sheet" if n == 1 else "multi-sheet"
        rel = (
            path.relative_to(base)
            if path.is_absolute() and str(path).startswith(str(base))
            else path
        )
        print(f"\n{rel}: ({mode})")
        for out, status, msg in convert_one(path, mode):
            marker = {"OK": "✅", "FAIL": "❌", "SKIP": "⏭️"}[status]
            rel_out = out.relative_to(base) if str(out).startswith(str(base)) else out
            print(f"  {marker} {rel_out} — {msg}")
            if status == "FAIL":
                all_ok = False

    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
